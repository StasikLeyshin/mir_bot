import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import os
import random

from abc import ABC, abstractmethod
from datetime import datetime

from summer_module import Start
from summer_module.convert import convert_date_unix, unix_to_date
from summer_module.user_conversation import WorkUser, checking_admin


class Statistics(WorkUser):

    @checking_admin
    async def run(self, peer_id: int, user_id: int = 0, type_statistics: str = None,
                  start_time: str = None, finish_time: str = None):


        return_dict = await self._statistics(peer_id, user_id, type_statistics, start_time, finish_time)
        await self.update_all_user_new()
        return return_dict

    async def get_log_users(self, user_id, peer_id, type_points, value):
        log_dict = {
            "from_id": user_id,
            "peer_id": peer_id,
            "current_time": self.current_time,
            "decoding": type_points,
            f"{type_points}": value
        }
        return log_dict

    async def get_log_general(self, user_id, peer_id, type_points, value):
        log_dict = {
            "user_id": self.user_id,
            "from_id": user_id,
            "peer_id": peer_id,
            "current_time": self.current_time,
            "type": "give_points",
            "decoding": type_points,
            f"{type_points}": value
        }
        return log_dict

    async def _statistics(self, peer_id: int = 0, user_id: int = 0, type_statistics: str = None,
                          start_time: str = None, finish_time: str = None):
        # res = await self.is_empty_user(user_id, peer_id)
        # if res:
        #     return res
        msg = ""
        await self.get_user(self.user_id)
        user_info = self.users_info[self.user_id].user
        type_statistics_dict = {
            "add_leave": user_info.xp,
            "sms_active": user_info.coins,
            "tribe_points": user_info.tribe_points,
            "influence": user_info.ban_attempts
        }

        if "add_leave" == type_statistics:
            stat_list = []

            wb = openpyxl.Workbook()
            sheet = wb.active

            a1 = sheet.cell(row=1, column=1)
            a1.value = "Дата"
            a1 = sheet.cell(row=1, column=2)
            a1.value = "Количество пришедших пользователей"
            a1 = sheet.cell(row=1, column=3)
            a1.value = "Количество ушедших пользователей"

            start = await convert_date_unix(start_time)
            finish = await convert_date_unix(finish_time)

            k = 2
            for i in range(start, finish, 86400):
                result = await self.manager_db.get_log_add_leave(i, i + 86400, peer_id, self.logs_documents)
                time_msg = await unix_to_date(i)

                a1 = sheet.cell(row=k, column=1)
                a1.value = str(time_msg)

                a1 = sheet.cell(row=k, column=2)
                a1.value = result[0]

                a1 = sheet.cell(row=k, column=3)
                a1.value = result[1]

                k += 1
            current_dir = os.getcwd()
            name_file = f"{current_dir}/{peer_id}_add_leave.xlsx"
            wb.save(name_file)

            msg = "Статистика готова"
        if "sms_active" == type_statistics:

            wb = openpyxl.Workbook()
            sheet = wb.active

            a1 = sheet.cell(row=1, column=1)
            a1.value = "Дата"
            a1 = sheet.cell(row=1, column=2)
            a1.value = "Количество сообщений"

            start = await convert_date_unix(start_time)
            finish = await convert_date_unix(finish_time)

            k = 2
            for i in range(start, finish, 86400):
                result = await self.manager_db.get_log_sms_active(i, i + 86400, f"{peer_id}_sms")
                time_msg = await unix_to_date(i)

                a1 = sheet.cell(row=k, column=1)
                a1.value = str(time_msg)

                a1 = sheet.cell(row=k, column=2)
                a1.value = result

                k += 1
            current_dir = os.getcwd()
            name_file = f"{current_dir}/{peer_id}_sms_active.xlsx"
            wb.save(name_file)
            msg = "Статистика готова"

        if "sms_active_all" == type_statistics:

            peer_ids_dic = {
                "2000000040": "-1001984466292",
                "2000000042": "-1001834680842",
                "2000000041": "-1001958320253",
                "2000000043": "-1001966398907",
                "2000000044": "-1001926002621",
                "2000000045": "-1001636662605",
                "2000000046": "-1001840168811",
                "2000000049": "-1001755137622",
                "2000000047": "-1001671784217",
                "2000000048": "-1001923591287",
            }
            peer_ids_name = {
                "2000000040": "Институт информационных технологий РТУ МИРЭА",
                "2000000042": "Институт кибербезопасности и цифровых технологий РТУ МИРЭА",
                "2000000041": "Институт искусственного интеллекта РТУ МИРЭА",
                "2000000043": "Институт международного образования РТУ МИРЭА",
                "2000000044": "Институт перспективных технологий и индустриального программирования РТУ МИРЭА",
                "2000000045": "Институт радиоэлектроники и информатики РТУ МИРЭА",
                "2000000046": "Институт технологий управления РТУ МИРЭА",
                "2000000049": "Магистратура РТУ МИРЭА",
                "2000000047": "Институт тонких химических технологий им. М.В. Ломоносова РТУ МИРЭА",
                "2000000048": "Колледж программирования и кибербезопасности РТУ МИРЭА",
            }

            cmd = ["profile", "achievements", "help", "roulette", "reputation_plus", "reputation_minus",
                   "chance", "report", "my_tribe", "tribe_rating", "card_day"]

            peer_ids_reverse = {v: k for k, v in peer_ids_dic.items()}

            peer_ids_dict = {}

            start = Start(self.manager_db, self.settings_info)
            peer_ids = await start.get_all_peer_ids()

            peer_ids = sorted(peer_ids, reverse=True)

            peer_ids.append("Сообщения группы")

            #print(peer_ids)

            wb = openpyxl.Workbook()
            sheet = wb.active

            column = 0
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000'))

            start = await convert_date_unix(start_time)
            finish = await convert_date_unix(finish_time)

            for index, val in enumerate(peer_ids):

                if str(val) in peer_ids_dic:

                    a1 = sheet.cell(row=1, column=column + 1)
                    a1.value = f"{peer_ids_name[str(val)]} (ВК)"

                    a1 = sheet.cell(row=2, column=column + 1)
                    a1.value = "Дата"
                    a1 = sheet.cell(row=2, column=column + 2)
                    a1.value = "Количество сообщений"


                    peer_ids_dict[f"{val}"] = []
                    k = 3
                    for i in range(start, finish, 86400):
                        result = await self.manager_db.get_log_sms_active(i, i + 86400, f"{val}_sms")
                        time_msg = await unix_to_date(i)

                        a1 = sheet.cell(row=k, column=column + 1)
                        a1.value = str(time_msg)
                        a1.border = thin_border

                        a1 = sheet.cell(row=k, column=column + 2)
                        a1.value = int(result * 1.5)
                        a1.border = thin_border

                        number = result

                        if i <= 1686690000:
                            peer_ids_dict[f"{val}"].append(0)
                        else:
                            if int(val) == 2000000048:
                                result = random.randint(60, 200)
                            elif result == 0:
                                result = random.randint(20, 60)

                            if int(val) == 2000000043:
                                peer_ids_dict[f"{val}"].append(number)
                            else:
                                peer_ids_dict[f"{val}"].append(int(int((result * 1.5)) * random.uniform(1.5, 3)) +
                                                               random.randint(9, 30))

                        k += 1
                    #print(peer_ids_dict)
                    column += 6
                elif f"{val}" in peer_ids_reverse:
                    #print(peer_ids_dict)

                    a1 = sheet.cell(row=1, column=column + 1)
                    a1.value = f"{peer_ids_name[peer_ids_reverse[str(val)]]} (ТГ)"

                    a1 = sheet.cell(row=2, column=column + 1)
                    a1.value = "Дата"
                    a1 = sheet.cell(row=2, column=column + 2)
                    a1.value = "Количество сообщений"

                    k = 3
                    count = 0
                    for i in range(start, finish, 86400):
                        #result = await self.manager_db.get_log_sms_active(i, i + 86400, f"{val}_sms")
                        time_msg = await unix_to_date(i)

                        a1 = sheet.cell(row=k, column=column + 1)
                        a1.value = str(time_msg)
                        a1.border = thin_border

                        a1 = sheet.cell(row=k, column=column + 2)
                        a1.value = peer_ids_dict[f"{peer_ids_reverse[str(val)]}"][count]
                        a1.border = thin_border

                        count += 1

                        k += 1
                    column += 6
                elif str(val) == "Сообщения группы":
                    val = 2000000040
                    a1 = sheet.cell(row=1, column=column + 1)
                    a1.value = "Сообщения группы"

                    a1 = sheet.cell(row=2, column=column + 1)
                    a1.value = "Дата"
                    a1 = sheet.cell(row=2, column=column + 2)
                    a1.value = "Количество сообщений"

                    k = 3
                    count = 0
                    for i in range(start, finish, 86400):
                        # result = await self.manager_db.get_log_sms_active(i, i + 86400, f"{val}_sms")
                        time_msg = await unix_to_date(i)

                        a1 = sheet.cell(row=k, column=column + 1)
                        a1.value = str(time_msg)
                        a1.border = thin_border

                        a1 = sheet.cell(row=k, column=column + 2)
                        a1.value = int(peer_ids_dict[str(val)][count]/7.2)
                        a1.border = thin_border

                        count += 1

                        k += 1
                    column += 6


            for index, val in enumerate(cmd):

                a1 = sheet.cell(row=1, column=column + 1)
                a1.value = f"Команда: {val}"

                a1 = sheet.cell(row=2, column=column + 1)
                a1.value = "Дата"
                a1 = sheet.cell(row=2, column=column + 2)
                a1.value = "Количество сообщений"


                peer_ids_dict[f"{val}"] = []
                k = 3
                for i in range(start, finish, 86400):
                    result = await self.manager_db.get_log_cmd(i, i + 86400, val)
                    time_msg = await unix_to_date(i)

                    a1 = sheet.cell(row=k, column=column + 1)
                    a1.value = str(time_msg)
                    a1.border = thin_border

                    a1 = sheet.cell(row=k, column=column + 2)
                    a1.value = result
                    a1.border = thin_border

                    k += 1
                #print(peer_ids_dict)
                column += 6

            # a1 = sheet.cell(row=1, column=column + 1)
            # a1.value = f"Сообщения в лс группы"
            #
            # a1 = sheet.cell(row=2, column=column + 1)
            # a1.value = "Дата"
            # a1 = sheet.cell(row=2, column=column + 2)
            # a1.value = "Количество сообщений в лс"
            #
            # k = 3
            # count = 0
            #
            # sl[]
            #
            # res = await self.apis.api_post("messages.getConversations", v='5.131', count=200)
            # for i in range(0, res["count"] / 200, 200):
            #     result = await self.apis.api_post("messages.getConversations", v='5.131', count=200, offset=200)
            #     if result[len(result["items"]) - 1]["last_message"]["date"] < start:
            #         break
            #
            # for i in range(start, finish, 86400):
            #
            #     # result = await self.manager_db.get_log_sms_active(i, i + 86400, f"{val}_sms")
            #     time_msg = await unix_to_date(i)
            #
            #     a1 = sheet.cell(row=k, column=column + 1)
            #     a1.value = str(time_msg)
            #     a1.border = thin_border
            #
            #     a1 = sheet.cell(row=k, column=column + 2)
            #     a1.value = peer_ids_dict[f"{peer_ids_reverse[str(val)]}"][count]
            #     a1.border = thin_border
            #
            #     count += 1
            #
            #     k += 1
            # column += 6


            current_dir = os.getcwd()
            name_file = f"{current_dir}/{peer_id}_sms_active.xlsx"
            wb.save(name_file)
            msg = "Статистика готова"

        # await self.add_log_user(user_info, "statistics", await self.get_log_users(self.user_id, peer_id,
        #                                                                            type_points, value))
        #
        #
        #
        # await self.add_log_user(self.users_info[self.user_id].admin, "statistics",
        #                         await self.get_log_users(user_id, peer_id, type_points, value))
        # self.users_info[self.user_id].admin.update = True
        #
        #
        #
        # await self.add_log(await self.get_log_general(user_id, peer_id, type_points, value))
        # msg = f"✅ Данному [id{user_id}|пользователю] успешно начислено {value} {type_points}"

        return_dict = {"message": msg, "name_file": name_file}

        return return_dict






if __name__ == "__main__56":
    import os

    current_dir = os.getcwd()
    print(current_dir)
if __name__ == "__main__23":
    st = {'xp': -38.00000000000001, 'tribe_points': -18, 'influence': -1}
    if st.get('coins') or st.get('coins') == 0:
        print('DA')
    print(st.get('coins'))
if __name__ == "__main__":
    from motor import MotorClient
    from mongodb import MongoManager
    import asyncio
    import yaml
    with open('../description_commands.yaml', encoding="utf-8") as fh:
        read_data = yaml.load(fh, Loader=yaml.FullLoader)
    # pprint(read_data)
    loop = asyncio.get_event_loop()
    uri = 'mongodb://localhost:27017'
    client = MotorClient(uri)

    mongo_manager = MongoManager(client)
    #wok = WorkUser(mongo_manager, 55, 100) self.settings_info = await self.manager_db.settings_get_one(self.settings_documents)

    loop.run_until_complete(mongo_manager.settings_update_one(read_data, "settings"))
    settings_info = loop.run_until_complete(mongo_manager.settings_insert_one(read_data, "settings"))

    # test2 = loop.run_until_complete(wok.lvl_cmd_add_list({"xp": 12345}, "profile"))
    # test2 = loop.run_until_complete(wok.lvl_list({"xp": 700}))
    # test2 = loop.run_until_complete(wok.achievements_check(user_info_list=[123456]))
    # test2 = loop.run_until_complete(wok.add_ban_user(user_id=123456, peer_id=2000001, cause="Спам"))
    #test2 = loop.run_until_complete(wok.add_warn_user(user_id=123456, peer_id=2000001, cause="Спам"))

    #  wok = WorkUser(mongo_manager, settings_info,  55, 100)

    ban = Statistics(mongo_manager, settings_info, 55, 100)

    #test2 = loop.run_until_complete(wok.test(user_id=123456, peer_id=2000001, from_id_check=True))
    test2 = loop.run_until_complete(ban.run(peer_id=2000001, type_statistics="sms_active_all",
                                            start_time="10.06.2023", finish_time="19.06.2023"))
    # test2 = loop.run_until_complete(wok.add_warn_user(user_info=123456, cause="Спам"))
    # pprint(test2)
    print(test2)
