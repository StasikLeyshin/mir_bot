
from PIL import ImageDraw, ImageFont, Image
import textwrap
import os
import ujson
import json
import re
from openpyxl import load_workbook


class generating:

    def __init__(self, file, create_mongo):
        self.file = file
        self.create_mongo = create_mongo

    async def sp_vopr(self, sheet, f=0, nap="bac"):


        # if encoding == 0:
        #     f2 = open(f"{self.file}.txt", "r+", encoding="cp1251")
        # else:
        #     f2 = open(f"{self.file}.txt", "r+", encoding="utf8")
        # s = f2.readlines()
        # f2.seek(0)
        # f2.close()
        # vopr = {"nom": {}, "vopr": {}, "spis_vop": {}, "att": " "}
        # vopr_new = []
        # for i in s:
        #     pas = i[i.find(":") + 1:]
        #     logi = i[:i.find(":")]
        #     nom = pas[pas.find(":") + 1:].replace("\\v", "\n")
        #     vopr["nom"][logi] = nom
        #     vopr["vopr"][pas[:pas.find(":")]] = pas[pas.find(":") + 1:].replace("\\v", "\n")
        #     vopr["spis_vop"][logi] = pas[:pas.find(":")]
        #     vopr_new.append({"nom": logi, "vopr": pas[:pas.find(":")], "otvet": nom})
        wb = load_workbook(f'{self.file}.xlsx')

        # Get sheet names
        #print(wb.get_sheet_names())
        sheet = wb.get_sheet_by_name(f'{sheet}')
        vopr = {"nom": {}, "vopr": {}, "spis_vop": {}, "att": " "}
        vopr_new = []
        for row in enumerate(sheet.rows):
            # print(row)
            if row[0] > 0:
                for cell in enumerate(row[1]):

                    if cell[1].value == None:
                        pass
                    else:
                        if cell[0] == 0:
                            q = cell[1].value
                            vopr["spis_vop"][str(row[0])] = cell[1].value
                        if cell[0] == 1:
                            otvet = cell[1].value
                            vopr["nom"][str(row[0])] = cell[1].value
                        # if cell[0] == 3:
                        #     vopr["vopr"][str(q)] = str(otvet)
                        #     vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})
                        # print(cell[1].value)
                if row[1][0].value:
                    #print(row[1][0])
                    # print(row[1][0])
                    vopr["vopr"][str(q)] = str(otvet)
                    vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})
        #print(vopr)
        #print(vopr_new)

        if f == 1:

            MAX_W, MAX_H = 1080, 1920
            file_name = "generating_questions/img/1.jpg"
            im1 = Image.open(file_name)
            draw1 = ImageDraw.Draw(im1)
            im2 = Image.open(file_name)
            draw2 = ImageDraw.Draw(im2)
            im3 = Image.open(file_name)
            draw3 = ImageDraw.Draw(im3)
            im4 = Image.open(file_name)
            draw4 = ImageDraw.Draw(im3)
            if len(vopr["spis_vop"]) > 63:
                im4 = Image.open(file_name)
                draw4 = ImageDraw.Draw(im4)
            # font = ImageFont.truetype( "mira.ttf", 45)
            font = ImageFont.truetype("generating_questions/fonts/PFSquareSansPro-Medium.ttf", 45)
            font1 = ImageFont.truetype("generating_questions/fonts/PFSquareSansPro-Regular.ttf", 45)
            current_h, pad = 50, 10
            # para = astr.split(" ")
            margin = offset = 90
            nom2 = 0
            nom3 = 0
            nom4 = 0
            for i in vopr["spis_vop"].keys():
                if int(i) < 25:
                    draw = draw1
                elif int(i) > 25 and nom2 == 0:
                    draw = draw2
                    margin = offset = 90
                    nom2 = 1
                elif int(i) > 42 and nom3 == 0:
                    draw = draw3
                    margin = offset = 90
                    nom3 = 1
                elif int(i) > 63 and nom4 == 0:
                    draw = draw4
                    margin = offset = 90
                    nom4 = 1
                text = vopr["spis_vop"][i]
                no = str(i) + ". "
                if len(no) == 4:
                    dob = 75
                    margin = 77
                else:
                    dob = 49
                    margin = 100
                if len(text) >= 40 and len(text) < 60:
                    peren = 28
                elif len(text) < 40:
                    peren = 39
                elif len(text) >= 60:
                    peren = 40
                f = 0
                for line in textwrap.wrap(text, width=peren):
                    # print(line)
                    if f == 0:
                        draw.text((margin, offset), no, font=font)
                        draw.text((margin + dob, offset), line, font=font1)
                        offset += font.getsize(line)[1] + 10
                        f = 1
                    else:
                        # margin += dob
                        draw.text((margin + dob, offset), line, font=font1)
                        offset += font.getsize(line)[1] + 10
            im1.save(f'generating_questions/img/{nap}_test1.png')
            if nom2 == 1:
                im2.save(f'generating_questions/img/{nap}_test2.png')
            if nom3:
                im3.save(f'generating_questions/img/{nap}_test3.png')
            if nom4 == 1:
                im4.save(f'generating_questions/img/{nap}_test4.png')
            # if nap == "bak":
            #     self.create_mongo.questions_update("bots", "questions_abitur", vopr_new)
            # elif nap == "col":
            #     self.create_mongo.questions_update("bots", "questions_col", vopr_new)
            # elif nap == "mag":
            #     self.create_mongo.questions_update("bots", "questions_mag", vopr_new)
        #print(vopr)
        #json.dump(vopr, open("vopr.txt", "w", encoding="utf-8"))
        self.create_mongo.questions_update("bots", f"{nap}", vopr_new)
        """att = str(await photo_uploader.upload_message_photo("test1.png")) + "," + str(
            await photo_uploader.upload_message_photo("test2.png")) + "," + str(
            await photo_uploader.upload_message_photo("test3.png"))
        os.remove("test1.png")
        os.remove("test2.png")
        os.remove("test3.png")
        if nom4 == 1:
            im4.save('test4.png')
            att = str(att) + "," + str(await photo_uploader.upload_message_photo("test4.png"))
            os.remove("test4.png")
        vopr["att"] = str(att)
        json.dump(vopr, open("vopr.txt", "w", encoding="utf-8"))"""
        return


class Tree_answer_new:
    def __init__(self, data, row, column,  parent=None, children=None, **kwargs):
        self.__dict__.update(kwargs)
        #self.name = name
        self.parent = parent
        # self.children = children
        # if children:
        #     self.children = children
        self.level = int(str(row) + str(column))

        self.data = data
        self.list_data = self.children_check()


        self.row = row
        self.column = column
        self.parent = parent
        self.children = []


    def children_check(self):
        # text = "Детский технопарк «Альтаир» — уникальный проект, созданный РТУ МИРЭА при поддержке Правительства Москвы." \
        #        "На площадке детского технопарка школьники могут принять участие в программах профориентации, " \
        #        "ознакомиться с работой высокотехнологичного оборудования и освоить программы дополнительного " \
        #        "образования, разработанные сотрудниками университета при участии индустриальных партнёров детского " \
        #        "технопарка. Что вы хотите узнать о Детском технопарке (укажите цифрой)? " \
        #        "\n1) Как записаться. \n2) Какие есть программы." \
        #        "\n3) С какого возраста можно прийти. \n4) Инфраструктура." \
        #        "\n5) Партнёры. \n6) Шаг назад. "
        #print(self.data)
        pattern = re.compile(r"(\d+)\)((\s\D+)+)")
        # for item in text:
        list_number = pattern.findall(self.data)
        list_number_new = []
        if len(list_number) == 0:
            list_number = [('1', 'Шаг назад')]
        #print(list_number)
        for i in range(len(list_number)):
            li = []
            for j in range(len(list_number[i])):
                text = list_number[i][j].replace("\n", " ").replace(".", "").strip()
                if len(text.encode('utf-8')) > 64:
                    li.append(str(text[:30] + "..."))
                else:
                    li.append(str(text))
            list_number_new.append(li)

        #print(list_number_new)
        return list_number_new

    def analysis(self, data, data_parent):
        if data.replace(".", "").lower() in data_parent.replace(".", "").lower():
            return False
        else:
            return True

    def insert(self, data, row, column, row_parent=None, column_parent=None, obj=None):
        """

        :param data:
        :param row_parent: номер строки родителя
        :param column_parent: номер столбца родителя
        :param row: номер строки ребёнка
        :param column: номер столбца ребёнка
        :param obj:
        :return:
        """
        #print(row, column, row_parent, column_parent, data)
        if "Возвращает пользователя к предыдущему" not in data:
            if row_parent and column_parent:

                if row_parent == self.row and column_parent == self.column:
                    #print("Dobavil ", self.row, self.column)
                    if self.analysis(data, self.data):

                        self.children.append(Tree_answer_new(data, row, column, self.__dict__))
                        return row, column
                    else:
                        return None

                else:
                    for i in reversed(self.children):
                        tt = i.insert(data, row, column, row_parent, column_parent)
                        if tt:
                            return tt
                    #return "First else"
            else:
                #if row >= self.row:
                #print(column, self.column, data)
                if column - 2 == self.column:
                    #print("DAAAA")
                    if self.analysis(data, self.data):
                        self.children.append(Tree_answer_new(data, row, column, self.__dict__))
                        return row, column
                    else:
                        return None
                else:
                    for i in reversed(self.children):
                        tt = i.insert(data, row, column)
                        if tt:
                            return tt
        return None
                #return "Second else"

    def print_tree_symmetric(self):
        """

        симметричный вывод дерева двоичного поиска
        """
        #print(self.row, self.column, self.data)
        #print()
        if len(self.children) != 0:pass
            #print(f"CHILDREN {len(self.children)}:\n\n")
        for i in self.children:
            i.print_tree_symmetric()
            #print(self.column, self.row, self.data)
        # if self.left:
        #     self.left.print_tree_symmetric()
        # if self.data:
        #     print(self.name, self.data)
        #
        # if self.left:
        #     if self.left.pointer:
        #         self.left.pointer.print_tree_symmetric()



    def search(self, number=None, level=None, text=None):
        #print(level, self.level)
        #if not number:
            #return self.data
        #print(len(self.children))
        if level == self.level:
            #print(text)
            if not number and not text:
                return self.level, self.data, self.list_data
            elif text:
                if "шаг назад" == text.lower().replace(".", "").strip():
                    return self.parent['level'], self.parent['data'], self.parent['list_data']
                for i in self.list_data:
                    #print(i[1].lower().strip(), text.lower().replace(".", "").strip())

                    if text.lower().replace(".", "").strip() == i[1].lower().replace(".", "").strip():
                        return self.children[int(i[0]) - 1].level, self.children[int(i[0]) - 1].data,\
                               self.children[int(i[0]) - 1].list_data
                #print(len(self.list_data), len(self.children), self.list_data)
            elif 0 <= number <= len(self.children):

                if len(self.list_data) == number and level != 11:
                    return self.parent['level'], self.parent['data'], self.parent['list_data']
                else:
                    return self.children[number-1].level, self.children[number-1].data,\
                           self.children[number-1].list_data

            elif len(self.list_data) == number:
                return self.parent['level'], self.parent['data'], self.parent['list_data']

        if level:
            for i in self.children:
                if text:
                    ii = i.search(text=text, level=level)
                else:
                    ii = i.search(number, level)
                if ii:
                    return ii
        else:
            return 1, self.data


class GeneratingCardTaro:
    def __init__(self, file):
        self.file = file

    def start(self, sheet):
        wb = load_workbook(f'{self.file}.xlsx')
        sheet = wb.get_sheet_by_name(f'{sheet}')
        card_days_dict = []
        for row in sheet.rows:
            name = ""
            text = ""
            for cell in row:
                if cell.value != None:
                    if cell.column == 1 and not cell.row == 1:
                        name = cell.value
                    elif cell.column == 3 and not cell.row == 1:
                        text = cell.value
            if name and text:
                card_days_dict.append({"file": name, "text": text})
        return card_days_dict

class generating_answers:

    def is_int(self, number):
        try:
            int(number)
            return True
        except:
            return False
    def __init__(self, file, create_mongo=None):
        self.file = file
        #self.create_mongo = create_mongo

    def start(self, sheet):
        wb = load_workbook(f'{self.file}.xlsx')

        # Get sheet names
        # print(wb.get_sheet_names())
        sheet = wb.get_sheet_by_name(f'{sheet}')
        vopr = {"nom": {}, "vopr": {}, "spis_vop": {}, "att": " "}
        vopr_new = []


        slov = {"text": "", "name": "root", "row": 1, "column": 1}

        # for row in enumerate(sheet.rows):
        #     for cell in enumerate(row):
        #         if not self.is_int(cell[1]):
        #             for i in
                    #print(row[0], cell[1][1].value)
        tree = None
        for row in sheet.rows:
            for cell in row:
                if cell.value != None:
                    #if cell
                    #slov["root"]
                    if cell.row == 1 and cell.column == 1:
                        tree = Tree_answer_new(cell.value, cell.row, cell.column)
                        row_parent = cell.row
                        column_parent = cell.column

                        slov["text"] = cell.value
                        slov[f"{cell.row} {cell.column}"] = {}

                    #print(cell.row, row_parent, cell.column, column_parent)
                    elif cell.row >= row_parent and cell.column - 2 == column_parent:
                        #print("Dobavil", cell.row, row_parent, cell.column, column_parent, cell.value)
                        tt = tree.insert(cell.value, cell.row, cell.column, row_parent, column_parent)
                        #print(tt)
                        if tt:
                            row_parent = tt[0]
                            column_parent = tt[1]
                        #tree.print_tree_symmetric()

                    else:
                        #break
                        #print(cell.row, cell.column, row_parent, column_parent, cell.value)
                        tt = tree.insert(cell.value, cell.row, cell.column)
                        if tt:
                            row_parent = tt[0]
                            column_parent = tt[1]
                        #print(row_parent, column_parent)
                    #print(cell.row, cell.column, cell, cell.value)
                    #print()
        #tree.print_tree_symmetric()

        return tree
        # level = 11
        # ff = tree.search(level=level)
        # print(ff[1])
        # while True:
        #     #ff = tree.search()
        #     #print(ff)
        #     #s = int(input("Введите число: "))
        #     s = input("Введите строку: ")
        #     ff = tree.search(text=s, level=level)
        #     level = ff[0]
        #     print(ff)

        # for row in sheet.iter_cols():
        #     for cell in row:
        #         print(cell.value, end="|")
        #     print("")


        # for row in enumerate(sheet.rows):
        #     if row[0] > 0:
        #         for cell in enumerate(row[1]):
        #             #print(cell)
        #             if cell[1].value == None:
        #                 pass
        #             else:
        #                 if cell[0] == 0:
        #                     q = cell[1].value
        #                     vopr["spis_vop"][str(row[0])] = cell[1].value
        #                 if cell[0] == 1:
        #                     otvet = cell[1].value
        #                     vopr["nom"][str(row[0])] = cell[1].value
        #                 # if cell[0] == 3:
        #                 #     vopr["vopr"][str(q)] = str(otvet)
        #                 #     vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})
        #                 # print(cell[1].value)
        #         if row[1][0].value:
        #             print(row[1][0].value)
        #             #print(row[1][0])
        #             # print(row[1][0])
        #             vopr["vopr"][str(q)] = str(otvet)
        #             vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})


if __name__ == "__main__":
    #generating_answers('test').start('Вопросы JS НОЯБРЬ')

    print(GeneratingCardTaro("ТАРО").start("Для бота"))

    #print('Подобрать мероприятие по интересам.'.encode('utf-8'))
    # text = "Детский технопарк «Альтаир» — уникальный проект, созданный РТУ МИРЭА при поддержке Правительства Москвы." \
    #        "На площадке детского технопарка школьники могут принять участие в программах профориентации, " \
    #        "ознакомиться с работой высокотехнологичного оборудования и освоить программы дополнительного " \
    #        "образования, разработанные сотрудниками университета при участии индустриальных партнёров детского " \
    #        "технопарка. Что вы хотите узнать о Детском технопарке (укажите цифрой)? " \
    #        "\n1) Как записаться \n2) Какие есть программы." \
    #        "\n3) С какого возраста можно прийти. \n4) Инфраструктура." \
    #        "\n5) Партнёры. \n6) Шаг назад. "
    # pattern = re.compile(r"(\d+)\)((\s\D+)+)\n")
    # # for item in text:
    # list_number = pattern.findall(text)
    # print(list_number)

    # txt = "Курсы подготовки к творческому вступительному испытанию"
    # print(txt[:37])

    # from flask import Flask
    # from flask import request
    # from flask_json import FlaskJSON, JsonError, json_response, as_json
    # from flask import jsonify
    #
    # app = Flask(__name__)
    #
    #
    # @app.route('/', methods=['GET', 'POST'])
    # def hello():
    #     data = request.get_json(force=True)
    #     print(data)
    #     return {"response": "Ответ от сервера 1"}
    #
    #
    # app.run()


    #print(test[9])

    # text = "Детский технопарк «Альтаир» — уникальный проект, созданный РТУ МИРЭА при поддержке Правительства Москвы." \
    #        "На площадке детского технопарка школьники могут принять участие в программах профориентации, " \
    #        "ознакомиться с работой высокотехнологичного оборудования и освоить программы дополнительного " \
    #        "образования, разработанные сотрудниками университета при участии индустриальных партнёров детского " \
    #        "технопарка. Что вы хотите узнать о Детском технопарке (укажите цифрой)? " \
    #        "\n1) Как записаться. \n2) Какие есть программы." \
    #        "\n3) С какого возраста можно прийти. \n4) Инфраструктура." \
    #        "\n5) Партнёры. \n6) Шаг назад. "
    #

    # text = "В РТУ МИРЭА реализуется ряд программ довузовской подготовки." \
    #        "Выберите, о каких программах вы хотели бы узнать (укажите цифрой)?" \
    #        "\n1) Показать все.\n2) Подготовка к ЕГЭ." \
    #        "\n3) Физ-мат школа.\n4) Подготовка к экзаменам магистратуры.\n" \
    #        "5) Курсы подготовки к творческому вступительному испытанию.\n6) Шаг назад."

    # text = "Детский технопарк «Альтаир» обладает современной инфраструктурой: он расположен на территории " \
    #        "МИРЭА — Российского технологического университета в отдельном здании площадью более 3,5 тысяч квадратных " \
    #        "метров. Лаборатории оснащены современным и дорогостоящим оборудованием — некоторые единицы разработаны " \
    #        "специально для детского технопарка или ввезены на территорию России в единичном экземпляре. Всё " \
    #        "оборудование максимально безопасно для проведения занятий именно со школьниками.Подробнее: " \
    #        "https://www.mirea.ru/education/children-s-technology-park-altair/about-technopark/ " \
    #        "Если вы хотите вернуться на предыдущий шаг, отправьте цифру 1. "
    #
    # pattern = re.compile(r"(\d+)\)((\s\D+)+)\.")
    # # # #for item in text:
    # print(pattern.findall(text))

    # class Meta(type):
    #     def __new__(mcs, name, bases, attr):
    #         prefix = attr.get("prefix")
    #         if prefix:
    #             attr[prefix + "_var1"] = "hello"
    #             attr[prefix + "_var2"] = "goodbye"
    #
    #         return type.__new__(mcs, name, bases, attr)
    #
    #
    # class myclass(object):
    #     #__metaclass__ = Meta
    #     prefix = "coffee"
    #
    #     def __init__(self):
    #         #self.__dict__.update(kwargs)
    #         self.tt = 1
    #         setattr(self.__class__, self.prefix + "_var1", "hello")
    #         setattr(self.__class__, self.prefix + "_var2", "goodbye")
    #         print("DA")
    #
    #     def mytest(self):
    #         print(self.coffee_var1)
    #         print(self.coffee_var2)
    #         print(self.__class__().tt)
    #         print(self.__dict__)
    #         return self
    #
    #
    # gg = myclass().mytest()







    # wb = load_workbook('FAQ.xlsx')
    #
    # # Get sheet names
    # print(wb.get_sheet_names())
    # sheet = wb.get_sheet_by_name('Магистратура')
    # vopr = {"nom": {}, "vopr": {}, "spis_vop": {}, "att": " "}
    # vopr_new = []
    # flag = False
    # for row in enumerate(sheet.rows):
    #     #print(row)
    #     if row[0] > 1:
    #         for cell in enumerate(row[1]):
    #
    #             if cell[1].value == None:
    #                 pass
    #             else:
    #                 if cell[0] == 0:
    #                     q = cell[1].value
    #                     vopr["spis_vop"][str(row[0])] = cell[1].value
    #                 if cell[0] == 1:
    #                     otvet = cell[1].value
    #                     vopr["nom"][str(row[0])] = cell[1].value
    #                 # if cell[0] == 3:
    #                 #     vopr["vopr"][str(q)] = str(otvet)
    #                 #     vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})
    #                 #print(cell[1].value)
    #         if row[1][0].value:
    #             #print(row[1][0])
    #             vopr["vopr"][str(q)] = str(otvet)
    #             vopr_new.append({"nom": str(row[0]), "vopr": str(q), "otvet": str(otvet)})
    # print(vopr)
    # print(vopr_new)
    # top_players = pd.read_excel('FAQ.xlsx')
    # top_players.head()

    # df1 = pd.read_excel(
    #     os.path.join('FAQ.xlsx'),
    #     engine='openpyxl', sheet_name="Колледж"
    # )
    # for i in df1.head():
    #     for j in
    #     print(i)
